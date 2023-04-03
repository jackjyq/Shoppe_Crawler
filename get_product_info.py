import os.path
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright
import playwright
from time import sleep
import re

excel_file = "product_info.xlsx"
fields = [
    "product_url",
    "product_name",
    "highest_price",
    "lowest_price",
    "image_url",
    "ratings",
]


def append_to_excel(excel_file: str, product_info: list) -> None:
    # create excel file when needed
    if not os.path.isfile(excel_file):
        workbook = Workbook()
        if worksheet := workbook.active:
            worksheet.append(fields)
        workbook.save(excel_file)
    # append product info to excel
    workbook = load_workbook(excel_file)
    if worksheet := workbook.active:
        worksheet.append(product_info)
    workbook.save(excel_file)


def parse_prices_text(prices_text: str) -> tuple[float, float]:
    """
    >>> parse_prices_text("฿85.1")
    (85.1, 85.1)
    >>> parse_prices_text("฿225 - ฿249")
    (225.0, 249.0)
    """
    regex = r"(\d+(\.\d+)?)"
    matches = re.findall(regex, prices_text)
    if len(matches) == 1:
        return (float(matches[0][0]), float(matches[0][0]))
    elif len(matches) == 2:
        return (float(matches[0][0]), float(matches[1][0]))


def get_product_info(url: str) -> list[str]:
    with sync_playwright() as p:
        cookies = [
            {"name": "language", "value": "en", "domain": "shopee.co.th", "path": "/"},
        ]
        product_info = fields.copy()

        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        context.add_cookies(cookies)
        page = context.new_page()

        print(f"Crawling product {url}")
        product_info[0] = url
        page.goto(url, wait_until="commit")
        page.wait_for_selector(".product-rating-overview__rating-score", timeout=100000)

        # 商品名
        if product_name_object := page.query_selector('meta[property="og:title"]'):
            product_name = product_name_object.get_attribute("content")
            product_info[1] = product_name
            print(product_name)
        # 商品首图 URL
        if image_object := page.query_selector('meta[property="og:image"]'):
            image_url = image_object.get_attribute("content")
            product_info[4] = image_url
            print(image_url)
        # 评分
        if rating_object := page.query_selector(
            ".product-rating-overview__rating-score"
        ):
            rating = rating_object.text_content()
            product_info[5] = rating
            print(rating)
        # 价格
        if prices_object := page.query_selector(".pqTWkA"):
            prices: tuple[float, float] = parse_prices_text(
                prices_object.text_content()
            )
            product_info[3] = prices[0]
            product_info[2] = prices[1]
        # if product_info_ojb := page.query_selector('script[type="application/ld+json"]:nth-of-type(2)'):
        #     product_info_json = product_info_ojb.text_content()
        #     product_info_dict = json.loads(product_info_json)
        #     print(product_info_dict)
        context.close()
        browser.close()
    return product_info


if __name__ == "__main__":
    with open("product_urls.txt", "r") as file:
        urls = file.readlines()
    for url in urls:
        try:
            product_info = get_product_info(url)
        except playwright._impl._api_types.TimeoutError:
            print(f'Crawl Error for {url}')
            with open('crawl_errors.txt', 'w') as file:
                file.write(url + "\n")
            continue

        append_to_excel(excel_file, product_info)
        sleep(5)

